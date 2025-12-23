from __future__ import annotations

from collections.abc import Iterable, Mapping
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

# Excel headers we accept for certificate numbers; include common typo + correct spelling.
CERTIFICATE_HEADER_KEYS: tuple[str, ...] = (
    "Номер свидетельтсва",
    "Номер свидетельства",
    "Свидетельство о поверке",
)


def extract_certificates_from_excel(file_bytes: bytes) -> list[str]:
    """Return certificate numbers from the first sheet matching known headers."""
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_cells:
            return []
        accepted = {header.lower() for header in CERTIFICATE_HEADER_KEYS}
        column_index: int | None = None
        for idx, cell_value in enumerate(header_cells):
            if isinstance(cell_value, str) and cell_value.strip().lower() in accepted:
                column_index = idx
                break
        if column_index is None:
            return []

        certs: list[str] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or column_index >= len(row):
                continue
            value = row[column_index]
            if value is None:
                continue
            text = str(value).strip()
            if text:
                certs.append(text)
        return certs
    finally:
        wb.close()


def read_column_as_list(
    file_bytes: bytes, column_letter: str = "P", start_row: int = 2
) -> list[str]:
    """
    Читает колонку (по умолчанию P) со start_row, убирает пустые и дубликаты (сохраняя порядок).
    """
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active
    items = []
    col = column_letter.upper()

    for row in range(start_row, ws.max_row + 1):
        val = ws[f"{col}{row}"].value
        if val is None:
            continue
        s = str(val).strip()
        if s:
            items.append(s)

    seen = set()
    uniq = []
    for x in items:
        if x not in seen:
            seen.add(x)
            uniq.append(x)
    return uniq


def _normalized_header(values: Iterable[str]) -> set[str]:
    return {str(value).strip().lower() for value in values if value}


def _extract_header(ws, header_row: int) -> list[str]:
    header_row = max(int(header_row or 1), 1)
    if header_row > ws.max_row:
        return []

    header: list[str] = []
    for cell in ws[header_row]:
        val = cell.value
        if val is None and ws.merged_cells.ranges:
            for merged in ws.merged_cells.ranges:
                if cell.coordinate in merged:
                    val = ws.cell(merged.min_row, merged.min_col).value
                    break
        if val is None or str(val).strip() == "":
            for offset in range(1, 3):
                probe_row = header_row - offset
                if probe_row < 1:
                    break
                probe_val = ws.cell(probe_row, cell.column).value
                if probe_val is not None and str(probe_val).strip():
                    val = probe_val
                    break
        header.append(str(val).strip() if val is not None else "")
    return header


def _sheet_rows(ws, header: list[str], start_row: int) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for r in ws.iter_rows(min_row=start_row, values_only=True):
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        item: dict[str, object] = {}
        for i, key in enumerate(header):
            if not key:
                continue
            if i < len(r):
                item[key] = r[i]
        rows.append(item)
    return rows


def read_rows_as_dicts(
    file_bytes: bytes,
    *,
    header_row: int = 1,
    data_start_row: int | None = None,
    sheet: str | None = None,
) -> list[dict[str, object]]:
    """
    Читает указанный лист (по умолчанию активный) как список словарей.

    :param header_row: Номер строки (1-based), содержащей заголовки.
    :param data_start_row: Первая строка с данными; по умолчанию header_row + 1.
    :param sheet: Имя листа. Если None — используется активный лист.
    Пустые строки пропускаются.
    """
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    try:
        ws = wb[sheet] if sheet else wb.active
        header = _extract_header(ws, header_row)
        start_row = int(data_start_row) if data_start_row else header_row + 1
        return _sheet_rows(ws, header, start_row)
    finally:
        wb.close()


def read_rows_with_required_headers(
    file_bytes: bytes,
    *,
    header_row: int = 1,
    data_start_row: int | None = None,
    required_headers: tuple[str, ...] | None = None,
) -> list[dict[str, object]]:
    """
    Перебирает все листы файла и возвращает строки первого листа,
    содержащего любой из required_headers (без учёта регистра/пробелов).
    Если подходящий лист не найден — возвращает пустой список.
    """

    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    try:
        wanted = _normalized_header(required_headers or [])
        for ws in wb.worksheets:
            header = _extract_header(ws, header_row)
            if wanted:
                normalized = _normalized_header(header)
                if not (normalized & wanted):
                    continue
            start_row = int(data_start_row) if data_start_row else header_row + 1
            rows = _sheet_rows(ws, header, start_row)
            if rows:
                return rows
        return []
    finally:
        wb.close()


def extract_certificate_number(row: Mapping[str, Any]) -> str:
    """Return trimmed certificate number from supported header variants."""

    normalized = {str(key).strip().lower(): value for key, value in row.items() if key}
    for header in CERTIFICATE_HEADER_KEYS:
        value = row.get(header)
        if value is None:
            value = normalized.get(header.lower())
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""
