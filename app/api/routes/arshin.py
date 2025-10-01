"""Routes for Arshin resolution endpoints."""

from __future__ import annotations

import asyncio
from typing import Any

import httpx
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile

from app.api.deps import get_http_client, get_semaphore
from app.services.arshin_client import (
    compose_etalon_line_from_details,
    extract_detail_fields,
    fetch_vri_details,
    fetch_vri_id_by_certificate,
    guess_year_from_cert,
)
from app.utils.excel import CERTIFICATE_HEADER_KEYS

router = APIRouter(prefix="/api/v1/resolve", tags=["arshin"])


@router.get("/vri-id")
async def get_vri_id(
    cert: str,
    client: httpx.AsyncClient = Depends(get_http_client),
    sem: asyncio.Semaphore = Depends(get_semaphore),
) -> dict[str, str | None]:
    """Возвращает vri_id по номеру свидетельства.

    - Проверяем входные данные
    - Определяем год из номера (если удаётся)
    - Ищем vri_id через Аршин
    """
    if not cert:
        raise HTTPException(status_code=400, detail="cert is required")

    year = guess_year_from_cert(cert)
    vri_id = await fetch_vri_id_by_certificate(
        client, cert, year=year, sem=sem, use_cache=False
    )
    if not vri_id:
        # возвращаем 200 с пустым vri_id — тестам это не критично,
        # но пусть будет информативнее
        return {"vri_id": None}
    return {"vri_id": vri_id}


@router.get("/vri/{vri_id}")
async def get_vri_details(
    vri_id: str,
    client: httpx.AsyncClient = Depends(get_http_client),
    sem: asyncio.Semaphore = Depends(get_semaphore),
) -> dict[str, Any]:
    """Возвращает краткую информацию по vri_id, включая строку эталона.

    - Проверяем входные данные
    - Запрашиваем детали по vri_id
    - Готовим строку эталона и ключевые поля
    """
    if not vri_id:
        raise HTTPException(status_code=400, detail="vri_id is required")

    details = await fetch_vri_details(client, vri_id, sem=sem)
    if not details:
        raise HTTPException(status_code=404, detail="not found")

    etalon_line = compose_etalon_line_from_details(details)
    fields = extract_detail_fields(details)
    return {
        "vri_id": vri_id,
        "etalon_line": etalon_line,
        "fields": fields,
        "result": details,  # на будущее
    }


# ───────────────────────── excel endpoint ─────────────────────────


def _read_cert_list_from_excel(file: UploadFile) -> list[str]:
    """Читает список сертификатов из Excel.

    - Ищет колонку по заголовку 'Номер свидетельства' (или частый вариант с опечаткой)
    - Возвращает непустые значения со 2-й строки и ниже
    """
    from openpyxl import load_workbook  # локальный импорт, чтобы не грузить при старте

    wb = load_workbook(file.file, read_only=True, data_only=True)
    ws = wb.active

    # Найти индекс колонки по заголовку
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_idx = None
    accepted_headers = {header.lower() for header in CERTIFICATE_HEADER_KEYS}
    for i, v in enumerate(header_row):
        if not isinstance(v, str):
            continue
        header = v.strip()
        if header and header.lower() in accepted_headers:
            col_idx = i
            break
    if col_idx is None:
        return []

    certs: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        val = row[col_idx] if col_idx < len(row) else None
        if val is None:
            continue
        cert = str(val).strip()
        if cert:
            certs.append(cert)
    return certs


@router.post("/vri-details-by-excel")
async def post_vri_details_by_excel(
    file: UploadFile = File(...),
    client: httpx.AsyncClient = Depends(get_http_client),
    sem: asyncio.Semaphore = Depends(get_semaphore),
) -> dict[str, Any]:
    """Принимает Excel и возвращает список найденных vri_id по сертификатам.

    Для каждого сертификата:
      - ищем vri_id (через /vri)
      - при наличии — подтягиваем детали (через /vri/{id})
    """
    certs = _read_cert_list_from_excel(file)

    # Обрабатываем строки параллельно, ограничение — через семафор
    async def process(cert: str) -> dict[str, Any]:
        try:
            vri_id = await fetch_vri_id_by_certificate(
                client, cert, year=guess_year_from_cert(cert), sem=sem, use_cache=False
            )
            details: dict[str, Any] = {}
            if vri_id:
                details = await fetch_vri_details(client, vri_id, sem=sem)
            return {"certificate": cert, "vri_id": vri_id, "details": details or None}
        except httpx.HTTPError as e:
            return {
                "certificate": cert,
                "vri_id": None,
                "details": None,
                "error": f"http error: {e}",
            }

    items = await asyncio.gather(*(process(c) for c in certs))

    return {"items": items}
