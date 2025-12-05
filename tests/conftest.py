import io
import os
import sys

import httpx
import pytest
from fastapi import FastAPI
from openpyxl import Workbook

# 1) добавляем корень репозитория в sys.path ДО импортов app.*
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

# 2) теперь можно импортировать код приложения
from app.main import app as fastapi_app  # noqa: E402
from app.utils.excel import CERTIFICATE_HEADER_KEYS  # noqa: E402


@pytest.fixture
def anyio_backend():
    """Force anyio to use asyncio backend to avoid pulling trio."""
    return "asyncio"


@pytest.fixture(scope="session")
def app() -> FastAPI:
    return fastapi_app


@pytest.fixture
async def async_client(app: FastAPI):
    transport = httpx.ASGITransport(app=app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as client:
        yield client


@pytest.fixture
def make_excel():
    """Фабрика: создаёт XLSX в памяти с данными в указанной колонке/строке."""

    def _make(
        values,
        column_letter="P",
        header: str = CERTIFICATE_HEADER_KEYS[-1],
        start_row=2,
    ):
        wb = Workbook()
        ws = wb.active
        ws[f"{column_letter}1"] = header
        r = start_row
        for v in values:
            ws[f"{column_letter}{r}"] = v
            r += 1
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    return _make


@pytest.fixture(autouse=True)
def _clear_global_caches():
    """Сбрасываем глобальные кеши между тестами, чтобы мокам не мешали сохранённые значения."""
    try:
        from app.services.cache import arshin_cache
        arshin_cache._data.clear()  # type: ignore[attr-defined]
    except Exception:
        pass

    try:
        from app.utils import signatures
        signatures._clear_caches_for_tests()
    except Exception:
        pass

    yield

    try:
        from app.services.cache import arshin_cache
        arshin_cache._data.clear()  # type: ignore[attr-defined]
    except Exception:
        pass

    try:
        from app.utils import signatures
        signatures._clear_caches_for_tests()
    except Exception:
        pass
