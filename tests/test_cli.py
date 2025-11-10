import argparse
from pathlib import Path

import pytest
from openpyxl import Workbook

from app import cli


class _DummySession:
    def __init__(self):
        self.commits = 0

    async def commit(self):
        self.commits += 1


class _DummySessionContext:
    def __init__(self, session: _DummySession):
        self._session = session

    async def __aenter__(self):
        return self._session

    async def __aexit__(self, exc_type, exc, tb):
        return False


class _DummySessionMaker:
    def __init__(self, session: _DummySession):
        self._session = session

    def __call__(self):
        return _DummySessionContext(self._session)


def _make_registry_file(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws["H5"] = "Заводской №/ Буквенно-цифровое обозначение"
    ws["L5"] = "Документ"
    ws["H6"] = "SN-001"
    ws["L6"] = "CERT-001"
    wb.save(path)


@pytest.mark.anyio
async def test_run_import_registry_happy_path(tmp_path, monkeypatch, capsys):
    registry_path = tmp_path / "registry.xlsx"
    _make_registry_file(registry_path)

    dummy_session = _DummySession()
    monkeypatch.setattr(cli, "get_sessionmaker", lambda: _DummySessionMaker(dummy_session))

    async def fake_ingest(session, **kwargs):
        assert kwargs["source_file"] == registry_path.name
        return {"processed": 1, "deactivated": 0}

    monkeypatch.setattr(cli, "ingest_registry_rows", fake_ingest)

    args = argparse.Namespace(
        file=str(registry_path),
        source_sheet=None,
        instrument_kind="controllers",
        header_row=5,
        data_start_row=6,
    )

    await cli._run_import_registry(args)

    out = capsys.readouterr().out
    assert "Imported 1 rows" in out
    assert dummy_session.commits == 1


@pytest.mark.anyio
async def test_run_import_registry_raises_for_bad_headers(tmp_path):
    registry_path = tmp_path / "registry.xlsx"
    wb = Workbook()
    wb.save(registry_path)

    args = argparse.Namespace(
        file=str(registry_path),
        source_sheet=None,
        instrument_kind=None,
        header_row=5,
        data_start_row=6,
    )

    with pytest.raises(ValueError):
        await cli._run_import_registry(args)


def test_cli_main_invokes_async_runner(monkeypatch):
    called = {}

    async def fake_run_import(args):
        called["args"] = args.file

    monkeypatch.setattr(cli, "_run_import_registry", fake_run_import)

    cli.main(["import-registry", "dummy.xlsx"])

    assert called["args"] == "dummy.xlsx"
