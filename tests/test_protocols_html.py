import io

import httpx
import pytest
import respx
from openpyxl import Workbook

from app.services.arshin_client import ARSHIN_BASE
from app.utils.excel import CERTIFICATE_HEADER_KEYS


def _make_excel_row(
    certificate: str,
    sn: str = "ABC123",
    date: str = "15.01.2025",
    header: str = CERTIFICATE_HEADER_KEYS[-1],
    details: str = "0…0,6 МПа",
    accuracy_class: str = "1,5",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = header
    ws["B1"] = "Заводской номер"
    ws["C1"] = "Дата поверки"
    ws["D1"] = "Прочие сведения"
    ws["E1"] = "КТ"

    ws["A2"] = certificate
    ws["B2"] = sn
    ws["C2"] = date
    ws["D2"] = details
    ws["E2"] = accuracy_class

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.anyio
@respx.mock
@pytest.mark.parametrize("header", CERTIFICATE_HEADER_KEYS)
async def test_html_by_excel_returns_html(async_client, header, monkeypatch):
    cert = "С-ВЯ/15-01-2025/402123271"
    vri_id = "1-XYZ"

    # 1) /vri (по номеру свидетельства)
    respx.get(f"{ARSHIN_BASE}/vri").mock(
        side_effect=[
            httpx.Response(200, json={"result": {"items": [{"vri_id": vri_id}]}}),
            httpx.Response(
                200,
                json={
                    "result": {
                        "items": [
                            {
                                "result_docnum": "ET-123",
                                "verification_date": "01.01.2025",
                                "valid_date": "31.12.2025",
                            }
                        ]
                    }
                },
            ),
        ]
    )

    # 2) /vri/{id}
    respx.get(f"{ARSHIN_BASE}/vri/{vri_id}").mock(
        return_value=httpx.Response(
            200,
            json={
                "result": {
                    "means": {
                        "mieta": [
                            {
                                "regNumber": "77090.19.1Р.00761951",
                                "mitypeNumber": "77090-19",
                                "mitypeTitle": "Преобразователи давления эталонные",
                                "notation": "ЭЛМЕТРО-Паскаль-04, Паскаль-04",
                                "manufactureNum": "3127",
                                "manufactureYear": 2020,
                                "rankCode": "1Р",
                                "rankTitle": "Эталон 1-го разряда",
                            }
                        ]
                    },
                    "vriInfo": {
                        "docTitle": "МИ 123-45",
                        "vrfDate": "15.01.2025",
                        "validDate": "14.01.2026",
                        "applicable": {"certNum": cert},
                    },
                }
            },
        )
    )

    async def fake_find_certs(client, details, sem=None, preferred_reg_numbers=None):
        return [
            {
                "docnum": "ET-123",
                "verification_date": "01.01.2025",
                "valid_date": "31.12.2025",
                "line": "свидетельство о поверке № ET-123; действительно до 31.12.2025г.",
                "reg_number": "77090.19.1Р.00761951",
                "manufacture_num": "3127",
                "mitype_number": "77090-19",
            }
        ]

    monkeypatch.setattr("app.api.routes.protocols.find_etalon_certificates", fake_find_certs)

    xlsx = _make_excel_row(cert, header=header)
    files = {
        "file": (
            "input.xlsx",
            xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    r = await async_client.post("/api/v1/protocols/html-by-excel", files=files)

    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/html")
    html = r.text
    assert "ПРОТОКОЛ ПЕРИОДИЧЕСКОЙ ПОВЕРКИ" in html
    assert "ABC123" in html  # serial number appears
    assert "МИ 123-45" in html  # methodology
    assert "77090-19" in html  # etalon line
    assert "Диапазон измерений:" in html
    assert "0…0,6 МПа" in html
    assert "Класс точности:" in html
    assert "1,5" in html
