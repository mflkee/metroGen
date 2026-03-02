import io
from pathlib import Path

import httpx
import pytest
import respx
from openpyxl import Workbook

from app.services.arshin_client import ARSHIN_BASE
from app.utils.excel import CERTIFICATE_HEADER_KEYS


def _make_protocols_excel_row(
    certificate: str,
    sn: str = "ABC123",
    date: str = "15.01.2025",
    header: str = CERTIFICATE_HEADER_KEYS[-1],
) -> bytes:
    """Готовит XLSX с минимально нужной шапкой для /protocols/context-by-excel."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = header
    ws["B1"] = "Заводской номер"
    ws["C1"] = "Дата поверки"

    ws["A2"] = certificate
    ws["B2"] = sn
    ws["C2"] = date

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_manometers_excel_row(
    *,
    certificate: str,
    serial: str,
    instrument_type: str = "13535-93",
    verifier: str = "Большаков С.Н.",
    date: str = "15.06.2025",
    pressure: str = "101,5 кПа",
    owner: str = 'ООО "РИ-ИНВЕСТ"',
    methodology: str = "МИ 2124-90",
    details: str = "(0 - 4) кгс/см²",
    non_etalon_refs: str = "",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Обозначение СИ"
    ws["B1"] = "Заводской номер"
    ws["E1"] = "Владелец СИ"
    ws["F1"] = "Дата поверки"
    ws["H1"] = "Методика поверки"
    ws["J1"] = "Прочие сведения"
    ws["I1"] = "СИ, применяемые при поверке (не эталоны)"
    ws["K1"] = "Поверитель"
    ws["M1"] = "Давление"
    ws["P1"] = "Свидетельство о поверке"

    ws["A2"] = instrument_type
    ws["B2"] = serial
    ws["E2"] = owner
    ws["F2"] = date
    ws["H2"] = methodology
    ws["J2"] = details
    ws["I2"] = non_etalon_refs
    ws["K2"] = verifier
    ws["M2"] = pressure
    ws["P2"] = certificate

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_manometers_db_excel(
    *,
    serial: str,
    certificate: str,
    protocol_number: str,
    verifier: str | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active

    ws["H5"] = "Заводской №/ Буквенно-цифровое обозначение"
    ws["L5"] = "Документ"
    ws["P5"] = "номер_протокола"
    ws["G5"] = "Дата поверки"
    ws["Q5"] = "Поверитель"

    ws["H6"] = serial
    ws["L6"] = certificate
    ws["P6"] = protocol_number
    ws["G6"] = "05.06.2025"
    if verifier is not None:
        ws["Q6"] = verifier

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_manometers_db_excel_multi(rows: list[dict[str, str]]) -> bytes:
    wb = Workbook()
    ws = wb.active

    ws["G5"] = "Дата поверки"
    ws["H5"] = "Заводской №/ Буквенно-цифровое обозначение"
    ws["L5"] = "Документ"
    ws["P5"] = "номер_протокола"
    ws["Q5"] = "Поверитель"

    for offset, payload in enumerate(rows, start=6):
        ws[f"G{offset}"] = payload.get("verification_date")
        ws[f"H{offset}"] = payload.get("serial")
        ws[f"L{offset}"] = payload.get("certificate")
        ws[f"P{offset}"] = payload.get("protocol_number")
        ws[f"Q{offset}"] = payload.get("verifier")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _patch_etalon_certificates(monkeypatch, *, docnum: str = "ET-123") -> None:
    async def fake_find_certs(client, details, sem=None):
        return [
            {
                "docnum": docnum,
                "verification_date": "01.01.2025",
                "valid_date": "31.12.2025",
                "line": f"свидетельство о поверке № {docnum}; действительно до 31.12.2025г.",
                "reg_number": "77090.19.1Р.00761951",
                "manufacture_num": "3127",
                "mitype_number": "77090-19",
            }
        ]

    monkeypatch.setattr("app.api.routes.protocols.find_etalon_certificates", fake_find_certs)


@pytest.mark.anyio
@respx.mock
@pytest.mark.parametrize("header", CERTIFICATE_HEADER_KEYS)
async def test_contexts_by_excel_happy_path(async_client, header, monkeypatch):
    cert = "С-ВЯ/15-01-2025/402123271"
    vri_id = "1-XYZ"

    # 1) /vri (по номеру свидетельства) → vri_id
    # 2) /vri (по данным эталона)      → свидетельство эталона (второй вызов того же URL)
    respx.get(f"{ARSHIN_BASE}/vri").mock(
        side_effect=[
            httpx.Response(200, json={"result": {"items": [{"vri_id": vri_id}]}}),
            httpx.Response(
                200,
                json={
                    "result": {
                        "items": [
                            {
                                "result_docnum": "ET-123",
                                "verification_date": "01.01.2025",
                                "valid_date": "31.12.2025",
                            }
                        ]
                    }
                },
            ),
        ]
    )

    # 3) /vri/{id} → детали для билдера
    respx.get(f"{ARSHIN_BASE}/vri/{vri_id}").mock(
        return_value=httpx.Response(
            200,
            json={
                "result": {
                    "means": {
                        "mieta": [
                            {
                                "regNumber": "77090.19.1Р.00761951",
                                "mitypeNumber": "77090-19",
                                "mitypeTitle": "Преобразователи давления эталонные",
                                "notation": "ЭЛМЕТРО-Паскаль-04, Паскаль-04",
                                "manufactureNum": "3127",
                                "manufactureYear": 2020,
                                "rankCode": "1Р",
                                "rankTitle": "Эталон 1-го разряда",
                            }
                        ]
                    },
                    "vriInfo": {
                        "docTitle": "МИ 123-45",
                        "vrfDate": "15.01.2025",
                        "validDate": "14.01.2026",
                        "applicable": {"certNum": cert},
                    },
                }
            },
        )
    )
    _patch_etalon_certificates(monkeypatch)

    xlsx = _make_protocols_excel_row(cert, header=header)
    files = {
        "file": (
            "input.xlsx",
            xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    r = await async_client.post("/api/v1/protocols/context-by-excel", files=files)

    assert r.status_code == 200
    body = r.json()
    assert len(body["items"]) == 1
    item = body["items"][0]

    # Проверяем основные поля
    assert item["certificate"] == cert
    assert item["vri_id"] == vri_id
    assert item["filename"] == "ABC123-б-150125-1"
    # В контексте должна появиться строка эталона
    assert "77090-19" in (item["context"] or {}).get("etalon_line", "")


@pytest.mark.anyio
@respx.mock
async def test_contexts_by_excel_empty_certificate(async_client):
    # Пустой номер → элемент с ошибкой
    xlsx = _make_protocols_excel_row("")
    files = {
        "file": (
            "input.xlsx",
            xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    r = await async_client.post("/api/v1/protocols/context-by-excel", files=files)

    assert r.status_code == 200
    body = r.json()
    assert len(body["items"]) == 1
    item = body["items"][0]
    assert item["error"] == "certificate number is empty"
    assert item["vri_id"] == ""


@pytest.mark.anyio
@respx.mock
async def test_contexts_by_excel_not_found(async_client):
    cert = "С-ВЯ/15-01-2025/000000001"
    # /vri не находит запись
    respx.get(f"{ARSHIN_BASE}/vri").mock(
        return_value=httpx.Response(200, json={"result": {"items": []}})
    )

    xlsx = _make_protocols_excel_row(cert)
    files = {
        "file": (
            "input.xlsx",
            xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    r = await async_client.post("/api/v1/protocols/context-by-excel", files=files)

    assert r.status_code == 200
    body = r.json()
    assert len(body["items"]) == 1
    item = body["items"][0]
    assert item["certificate"] == cert
    assert item["error"] == "not found"
    assert item["vri_id"] == ""


@pytest.mark.anyio
@respx.mock
async def test_manometers_pdf_files_happy_path(async_client, tmp_path, monkeypatch):
    cert = "С-ЕЖБ/05-06-2025/443771099"
    serial = "03607"
    protocol_num = "06/001/25"
    vri_id = "1-MANO"
    run_id = "run-aaa"
    excel_verifier = "Excel Verifier"
    db_verifier = "DB Verifier"

    manometers_xlsx = _make_manometers_excel_row(
        certificate=cert,
        serial=serial,
        verifier=excel_verifier,
        non_etalon_refs="71394-18/96320, 44154-16/419433, 77777-77/ABC777",
    )
    db_xlsx = _make_manometers_db_excel(
        serial=serial,
        certificate=cert,
        protocol_number=protocol_num,
        verifier=db_verifier,
    )

    async def fake_pdf(html: str) -> bytes | None:
        assert "ПРОТОКОЛ" in html
        assert "Измеритель влажности и температуры" in html
        assert "Секундомер электронный" in html
        assert "77777-77" in html
        assert "ABC777" in html
        assert "СИ, применяемые при поверке (не эталоны):" not in html
        return b"%PDF-manometer%"

    monkeypatch.setattr("app.api.routes.protocols.html_to_pdf_bytes", fake_pdf)
    monkeypatch.setattr("app.api.routes.protocols._make_run_id", lambda: run_id)
    _patch_etalon_certificates(monkeypatch)

    def _fake_named_exports_dir(name: str) -> Path:
        path = tmp_path / name
        path.mkdir(parents=True, exist_ok=True)
        return path

    monkeypatch.setattr("app.api.routes.protocols.get_named_exports_dir", _fake_named_exports_dir)

    respx.get(f"{ARSHIN_BASE}/vri").mock(
        side_effect=[
            httpx.Response(200, json={"result": {"items": [{"vri_id": vri_id}]}}),
            httpx.Response(
                200,
                json={
                    "result": {
                        "items": [
                            {
                                "result_docnum": "ET-123",
                                "verification_date": "01.01.2025",
                                "valid_date": "31.12.2025",
                            }
                        ]
                    }
                },
            ),
        ]
    )

    respx.get(f"{ARSHIN_BASE}/vri/{vri_id}").mock(
        return_value=httpx.Response(
            200,
            json={
                "result": {
                    "means": {
                        "mieta": [
                            {
                                "regNumber": "77090.19.1Р.00761951",
                                "mitypeNumber": "77090-19",
                                "mitypeTitle": "Преобразователи давления эталонные",
                                "notation": "ЭЛМЕТРО-Паскаль-04, Паскаль-04",
                                "manufactureNum": "3127",
                                "manufactureYear": 2020,
                                "rankCode": "1Р",
                                "rankTitle": "Эталон 1-го разряда",
                            }
                        ]
                    },
                    "vriInfo": {
                        "docTitle": "МИ 123-45",
                        "vrfDate": "15.06.2025",
                        "validDate": "14.06.2026",
                        "applicable": {"certNum": cert},
                    },
                }
            },
        )
    )

    files = {
        "manometers_file": (
            "manometers.xlsx",
            manometers_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        "db_file": (
            "09 БД.xlsx",
            db_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }

    aux_payloads = [
        {
            "title": "Измеритель влажности и температуры",
            "reg_number": "71394-18",
            "modification": "ИВТМ-7",
            "manufacture_num": "96320",
            "certificate_no": "С-ВСА/02-06-2025/436974158",
            "verification_date": "02.06.2025",
            "valid_to": "01.06.2026",
        },
        {
            "title": "Секундомер электронный",
            "reg_number": "44154-16",
            "modification": "Интеграл С-01",
            "manufacture_num": "419433",
            "certificate_no": "С-ВЯ/19-12-2024/397249365",
            "verification_date": "19.12.2024",
            "valid_to": "18.12.2025",
        },
    ]
    for payload in aux_payloads:
        create_aux = await async_client.post("/api/v1/auxiliary-instruments", json=payload)
        assert create_aux.status_code == 201

    response = await async_client.post("/api/v1/protocols/manometers/pdf-files", files=files)

    assert response.status_code == 200
    body = response.json()
    assert body["count"] == 1
    assert body["errors"] == []
    assert len(body["files"]) == 1

    saved_path = Path(body["files"][0])
    assert saved_path.exists()
    assert saved_path.read_bytes() == b"%PDF-manometer%"
    assert saved_path.parent == tmp_path / f"Generation pressure 09 - {run_id}"
    assert saved_path.name == "2025-06-15 № 03607 (МПИ-1).pdf"


@pytest.mark.anyio
@respx.mock
async def test_manometers_pdf_files_certificate_mismatch(async_client, tmp_path, monkeypatch):
    serial = "03607"
    excel_cert = "С-ЕЖБ/05-06-2025/443771999"
    db_cert = "С-ЕЖБ/05-06-2025/443771099"
    protocol_num = "06/001/25"

    manometers_xlsx = _make_manometers_excel_row(certificate=excel_cert, serial=serial)
    db_xlsx = _make_manometers_db_excel(
        serial=serial, certificate=db_cert, protocol_number=protocol_num
    )

    calls = {"pdf": 0}

    async def fake_pdf(html: str) -> bytes | None:
        calls["pdf"] += 1
        return b"%PDF%"

    monkeypatch.setattr("app.api.routes.protocols.html_to_pdf_bytes", fake_pdf)

    def _fake_named_exports_dir(name: str) -> Path:
        path = tmp_path / name
        path.mkdir(parents=True, exist_ok=True)
        return path

    monkeypatch.setattr("app.api.routes.protocols.get_named_exports_dir", _fake_named_exports_dir)

    files = {
        "manometers_file": (
            "manometers.xlsx",
            manometers_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        "db_file": (
            "09 БД.xlsx",
            db_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }

    response = await async_client.post("/api/v1/protocols/manometers/pdf-files", files=files)

    assert response.status_code == 200
    body = response.json()
    assert body["count"] == 0
    assert body["files"] == []
    assert len(body["errors"]) == 1
    error = body["errors"][0]
    assert error["reason"] == "certificate mismatch between excel and db"
    assert error["serial"] == serial
    assert calls["pdf"] == 0


@pytest.mark.anyio
@respx.mock
async def test_manometers_pdf_files_prefers_matching_month(async_client, tmp_path, monkeypatch):
    serial = "03607"
    july_cert = "С-ЕЖБ/05-07-2025/443771777"
    aug_cert = "С-ЕЖБ/05-08-2025/443771888"
    july_protocol = "07/001/25"
    aug_protocol = "08/001/25"

    manometers_xlsx = _make_manometers_excel_row(
        certificate=july_cert, serial=serial, date="15.07.2025"
    )
    db_xlsx = _make_manometers_db_excel_multi(
        [
            {
                "serial": serial,
                "certificate": aug_cert,
                "protocol_number": aug_protocol,
                "verification_date": "05.08.2025",
            },
            {
                "serial": serial,
                "certificate": july_cert,
                "protocol_number": july_protocol,
                "verification_date": "05.07.2025",
            },
        ]
    )

    async def fake_pdf(html: str) -> bytes | None:
        return b"%PDF-month%"

    monkeypatch.setattr("app.api.routes.protocols.html_to_pdf_bytes", fake_pdf)
    _patch_etalon_certificates(monkeypatch)

    def _fake_named_exports_dir(name: str) -> Path:
        path = tmp_path / name
        path.mkdir(parents=True, exist_ok=True)
        return path

    monkeypatch.setattr("app.api.routes.protocols.get_named_exports_dir", _fake_named_exports_dir)

    vri_id = "1-MONTH"
    respx.get(f"{ARSHIN_BASE}/vri").mock(
        side_effect=[
            httpx.Response(200, json={"result": {"items": [{"vri_id": vri_id}]}}),
            httpx.Response(
                200,
                json={
                    "result": {
                        "items": [
                            {
                                "result_docnum": "ET-123",
                                "verification_date": "01.07.2025",
                                "valid_date": "31.12.2025",
                            }
                        ]
                    }
                },
            ),
        ]
    )

    respx.get(f"{ARSHIN_BASE}/vri/{vri_id}").mock(
        return_value=httpx.Response(
            200,
            json={
                "result": {
                    "means": {
                        "mieta": [
                            {
                                "regNumber": "77090.19.1Р.00761951",
                                "mitypeNumber": "77090-19",
                                "mitypeTitle": "Преобразователи давления эталонные",
                                "notation": "ЭЛМЕТРО-Паскаль-04, Паскаль-04",
                                "manufactureNum": "3127",
                                "manufactureYear": 2020,
                                "rankCode": "1Р",
                                "rankTitle": "Эталон 1-го разряда",
                            }
                        ]
                    },
                    "vriInfo": {
                        "docTitle": "МИ 123-45",
                        "vrfDate": "15.07.2025",
                        "validDate": "14.07.2026",
                        "applicable": {"certNum": july_cert},
                    },
                }
            },
        )
    )

    files = {
        "manometers_file": (
            "manometers.xlsx",
            manometers_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        "db_file": (
            "07 БД.xlsx",
            db_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }

    response = await async_client.post("/api/v1/protocols/manometers/pdf-files", files=files)
    assert response.status_code == 200

    body = response.json()
    assert body["count"] == 1
    assert body["errors"] == []

    cert_calls = [
        call
        for call in respx.calls
        if call.request.url.path.endswith("/vri")
        and call.request.url.params.get("result_docnum")
    ]
    assert any(
        call.request.url.params.get("result_docnum") == july_cert for call in cert_calls
    ), "expected lookup using July certificate"


@pytest.mark.anyio
@respx.mock
async def test_controllers_pdf_files_fail_when_pdf_unavailable(async_client, monkeypatch):
    controllers_xlsx = _make_manometers_excel_row(certificate="C-1", serial="001")
    db_xlsx = _make_manometers_db_excel(serial="001", certificate="C-1", protocol_number="01/25")

    async def fake_pdf(*_args, **_kwargs):
        return None

    monkeypatch.setattr("app.api.routes.protocols.html_to_pdf_bytes", fake_pdf)

    respx.get(f"{ARSHIN_BASE}/vri").mock(
        side_effect=[
            httpx.Response(200, json={"result": {"items": [{"vri_id": "CTRL-1"}]}}),
            httpx.Response(200, json={"result": {"items": []}}),
        ]
    )
    respx.get(f"{ARSHIN_BASE}/vri/CTRL-1").mock(
        return_value=httpx.Response(200, json={"result": {"means": {"mieta": []}, "vriInfo": {}}})
    )

    files = {
        "controllers_file": (
            "controllers.xlsx",
            controllers_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        "db_file": (
            "db.xlsx",
            db_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }

    response = await async_client.post("/api/v1/protocols/controllers/pdf-files", files=files)
    assert response.status_code == 500
    assert response.json()["detail"].startswith("PDF generation is unavailable")


@pytest.mark.anyio
@respx.mock
async def test_manometers_pdf_files_with_preloaded_registry(async_client, tmp_path, monkeypatch):
    cert = "С-ЕЖБ/05-06-2025/443771099"
    serial = "03607"
    protocol_num = "06/001/25"
    vri_id = "1-MANO"
    run_id = "run-bbb"

    manometers_xlsx = _make_manometers_excel_row(certificate=cert, serial=serial)
    db_xlsx = _make_manometers_db_excel(
        serial=serial, certificate=cert, protocol_number=protocol_num
    )

    async def fake_pdf(html: str) -> bytes | None:
        return b"%PDF-manometer%"

    monkeypatch.setattr("app.api.routes.protocols.html_to_pdf_bytes", fake_pdf)
    monkeypatch.setattr("app.api.routes.protocols._make_run_id", lambda: run_id)
    _patch_etalon_certificates(monkeypatch)

    def _fake_named_exports_dir(name: str) -> Path:
        path = tmp_path / name
        path.mkdir(parents=True, exist_ok=True)
        return path

    monkeypatch.setattr("app.api.routes.protocols.get_named_exports_dir", _fake_named_exports_dir)

    respx.get(f"{ARSHIN_BASE}/vri").mock(
        side_effect=[
            httpx.Response(200, json={"result": {"items": [{"vri_id": vri_id}]}}),
            httpx.Response(
                200,
                json={
                    "result": {
                        "items": [
                            {
                                "result_docnum": "ET-123",
                                "verification_date": "01.01.2025",
                                "valid_date": "31.12.2025",
                            }
                        ]
                    }
                },
            ),
        ]
    )

    respx.get(f"{ARSHIN_BASE}/vri/{vri_id}").mock(
        return_value=httpx.Response(
            200,
            json={
                "result": {
                    "means": {
                        "mieta": [
                            {
                                "regNumber": "77090.19.1Р.00761951",
                                "mitypeNumber": "77090-19",
                                "mitypeTitle": "Преобразователи давления эталонные",
                                "notation": "ЭЛМЕТРО-Паскаль-04, Паскаль-04",
                                "manufactureNum": "3127",
                                "manufactureYear": 2020,
                                "rankCode": "1Р",
                                "rankTitle": "Эталон 1-го разряда",
                            }
                        ]
                    },
                    "vriInfo": {
                        "docTitle": "МИ 123-45",
                        "vrfDate": "15.06.2025",
                        "validDate": "14.06.2026",
                        "applicable": {"certNum": cert},
                    },
                }
            },
        )
    )

    import_response = await async_client.post(
        "/api/v1/registry/import",
        files={
            "db_file": (
                "06 БД.xlsx",
                db_xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert import_response.status_code == 200

    response = await async_client.post(
        "/api/v1/protocols/manometers/pdf-files",
        files={
            "manometers_file": (
                "manometers.xlsx",
                manometers_xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["count"] == 1
    assert body["errors"] == []
    saved_path = Path(body["files"][0])
    assert saved_path.exists()
    assert saved_path.read_bytes() == b"%PDF-manometer%"
    assert saved_path.parent == tmp_path / f"Generation pressure 06 - {run_id}"
    assert saved_path.name == "2025-06-15 № 03607 (МПИ-1).pdf"


@pytest.mark.anyio
@respx.mock
async def test_contexts_by_excel_includes_multiple_etalons(async_client, monkeypatch):
    cert = "С-ВЯ/15-01-2025/402123777"
    vri_id = "1-MULTI"

    respx.get(f"{ARSHIN_BASE}/vri").mock(
        return_value=httpx.Response(200, json={"result": {"items": [{"vri_id": vri_id}]}})
    )

    respx.get(f"{ARSHIN_BASE}/vri/{vri_id}").mock(
        return_value=httpx.Response(
            200,
            json={
                "result": {
                    "means": {
                        "mieta": [
                            {
                                "regNumber": "77090.19.1Р.000111",
                                "mitypeNumber": "77090-19",
                                "mitypeTitle": "Эталон давления №1",
                                "notation": "ЭЛМЕТРО-Паскаль-04",
                                "manufactureNum": "E-001",
                            },
                            {
                                "regNumber": "77090.19.1Р.000222",
                                "mitypeNumber": "77090-19",
                                "mitypeTitle": "Эталон давления №2",
                                "notation": "ЭЛМЕТРО-Паскаль-05",
                                "manufactureNum": "E-002",
                            },
                        ]
                    },
                    "vriInfo": {
                        "docTitle": "МИ 123-45",
                        "vrfDate": "15.01.2025",
                        "validDate": "14.01.2026",
                        "applicable": {"certNum": cert},
                    },
                }
            },
        )
    )

    async def fake_find_certs(client, details, sem=None):
        return [
            {
                "line": "свидетельство о поверке № ET-001; действительно до 31.12.2025;",
                "manufacture_num": "E-001",
                "reg_number": "77090.19.1Р.000111",
            },
            {
                "line": "свидетельство о поверке № ET-002; действительно до 31.12.2025;",
                "manufacture_num": "E-002",
                "reg_number": "77090.19.1Р.000222",
            },
        ]

    monkeypatch.setattr("app.api.routes.protocols.find_etalon_certificates", fake_find_certs)

    xlsx = _make_protocols_excel_row(cert)
    files = {
        "file": (
            "input.xlsx",
            xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    response = await async_client.post("/api/v1/protocols/context-by-excel", files=files)

    assert response.status_code == 200
    ctx = response.json()["items"][0]["context"]
    entries = ctx["etalon_entries"]
    assert len(entries) == 2
    assert entries[0]["reg_number"].endswith("000111")
    assert entries[1]["reg_number"].endswith("000222")
    assert entries[1]["certificate_line"].startswith("свидетельство о поверке № ET-002")


@pytest.mark.anyio
@respx.mock
async def test_thermometers_pdf_files_happy_path(async_client, tmp_path, monkeypatch):
    cert = "С-ЕЖБ/05-10-2025/443771555"
    serial = "RTD-001"
    protocol_num = "10/001/25"
    vri_id = "1-RTD"
    run_id = "run-rtd"

    thermometers_xlsx = _make_manometers_excel_row(
        certificate=cert,
        serial=serial,
        instrument_type="71040-18",
        methodology="ГОСТ 8.461-2009",
        details="(0 - 100) °C",
        pressure="99,0 кПа",
    )
    db_xlsx = _make_manometers_db_excel(
        serial=serial,
        certificate=cert,
        protocol_number=protocol_num,
        verifier="DB RTD Verifier",
    )

    async def fake_pdf(html: str) -> bytes | None:
        return b"%PDF-rtd%"

    monkeypatch.setattr("app.api.routes.protocols.html_to_pdf_bytes", fake_pdf)
    monkeypatch.setattr("app.api.routes.protocols._make_run_id", lambda: run_id)
    _patch_etalon_certificates(monkeypatch, docnum="ET-RTD")

    def _fake_named_exports_dir(name: str) -> Path:
        path = tmp_path / name
        path.mkdir(parents=True, exist_ok=True)
        return path

    monkeypatch.setattr("app.api.routes.protocols.get_named_exports_dir", _fake_named_exports_dir)

    respx.get(f"{ARSHIN_BASE}/vri").mock(
        return_value=httpx.Response(200, json={"result": {"items": [{"vri_id": vri_id}]}})
    )
    respx.get(f"{ARSHIN_BASE}/vri/{vri_id}").mock(
        return_value=httpx.Response(
            200,
            json={
                "result": {
                    "means": {
                        "mieta": [
                            {
                                "regNumber": "71040.18.1Р.00761951",
                                "mitypeNumber": "71040-18",
                                "mitypeTitle": "Термопреобразователи сопротивления",
                                "notation": "ТСП-100",
                                "manufactureNum": "3127",
                                "manufactureYear": 2020,
                            }
                        ]
                    },
                    "vriInfo": {
                        "docTitle": "ГОСТ 8.461-2009",
                        "vrfDate": "15.10.2025",
                        "validDate": "14.10.2026",
                        "applicable": {"certNum": cert},
                    },
                }
            },
        )
    )

    files = {
        "thermometers_file": (
            "thermometers.xlsx",
            thermometers_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        "db_file": (
            "10 БД.xlsx",
            db_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }

    response = await async_client.post("/api/v1/protocols/thermometers/pdf-files", files=files)

    assert response.status_code == 200
    body = response.json()
    assert body["count"] == 1
    assert body["errors"] == []
    saved_path = Path(body["files"][0])
    assert saved_path.exists()
    assert saved_path.read_bytes() == b"%PDF-rtd%"
    assert saved_path.parent == tmp_path / f"Generation rtd 10 - {run_id}"


@pytest.mark.anyio
@respx.mock
async def test_manometers_failed_pdf_files_marks_failed_context(
    async_client, tmp_path, monkeypatch
):
    cert = "С-ЕЖБ/05-06-2025/443771099"
    serial = "03607"
    protocol_num = "06/001/25"
    vri_id = "1-MANO-FAILED"
    run_id = "run-failed"
    seen_contexts: list[dict[str, object]] = []

    manometers_xlsx = _make_manometers_excel_row(certificate=cert, serial=serial)
    db_xlsx = _make_manometers_db_excel(
        serial=serial,
        certificate=cert,
        protocol_number=protocol_num,
    )

    def fake_render(ctx: dict[str, object]) -> str:
        seen_contexts.append(dict(ctx))
        return "<html><body>FAILED</body></html>"

    async def fake_pdf(html: str) -> bytes | None:
        return b"%PDF-failed%"

    monkeypatch.setattr("app.api.routes.protocols.render_protocol_html", fake_render)
    monkeypatch.setattr("app.api.routes.protocols.html_to_pdf_bytes", fake_pdf)
    monkeypatch.setattr("app.api.routes.protocols._make_run_id", lambda: run_id)
    _patch_etalon_certificates(monkeypatch)

    def _fake_named_exports_dir(name: str) -> Path:
        path = tmp_path / name
        path.mkdir(parents=True, exist_ok=True)
        return path

    monkeypatch.setattr("app.api.routes.protocols.get_named_exports_dir", _fake_named_exports_dir)

    respx.get(f"{ARSHIN_BASE}/vri").mock(
        return_value=httpx.Response(200, json={"result": {"items": [{"vri_id": vri_id}]}})
    )
    respx.get(f"{ARSHIN_BASE}/vri/{vri_id}").mock(
        return_value=httpx.Response(
            200,
            json={
                "result": {
                    "means": {
                        "mieta": [
                            {
                                "regNumber": "77090.19.1Р.00761951",
                                "mitypeNumber": "77090-19",
                                "mitypeTitle": "Преобразователи давления эталонные",
                                "notation": "ЭЛМЕТРО-Паскаль-04, Паскаль-04",
                                "manufactureNum": "3127",
                                "manufactureYear": 2020,
                            }
                        ]
                    },
                    "vriInfo": {
                        "docTitle": "МИ 123-45",
                        "vrfDate": "15.06.2025",
                        "validDate": "14.06.2026",
                        "applicable": {"certNum": cert},
                    },
                }
            },
        )
    )

    files = {
        "manometers_file": (
            "manometers.xlsx",
            manometers_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
        "db_file": (
            "06 БД.xlsx",
            db_xlsx,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }

    response = await async_client.post("/api/v1/protocols/manometers/failed/pdf-files", files=files)

    assert response.status_code == 200
    body = response.json()
    assert body["count"] == 1
    assert body["errors"] == []
    saved_path = Path(body["files"][0])
    assert saved_path.exists()
    assert saved_path.read_bytes() == b"%PDF-failed%"
    assert saved_path.parent == tmp_path / f"Generation pressure failed 06 - {run_id}"

    assert seen_contexts
    assert seen_contexts[0].get("verification_failed") is True
    assert seen_contexts[0].get("hide_results_table") is True
    assert seen_contexts[0].get("conclusion_text") == "не годен"
    assert seen_contexts[0].get("table_rows") == []
