import io

import pytest
from openpyxl import Workbook

from app.utils.excel import (
    CERTIFICATE_HEADER_KEYS,
    extract_certificate_number,
    extract_certificates_from_excel,
    read_column_as_list,
)


def test_read_column_as_list_default():
    wb = Workbook()
    ws = wb.active
    ws["P1"] = CERTIFICATE_HEADER_KEYS[-1]
    ws["P2"] = "С-ЕЖБ/05-06-2025/440144576"
    ws["P3"] = "С-ЕЖБ/05-06-2025/440144575"
    ws["P4"] = "С-ЕЖБ/05-06-2025/440144575"  # дубликат
    buf = io.BytesIO()
    wb.save(buf)

    items = read_column_as_list(buf.getvalue(), column_letter="P", start_row=2)
    assert items == [
        "С-ЕЖБ/05-06-2025/440144576",
        "С-ЕЖБ/05-06-2025/440144575",
    ]


@pytest.mark.parametrize("header", CERTIFICATE_HEADER_KEYS)
def test_extract_certificate_number(header):
    row = {header: " С-ВЯ/15-01-2025/402123271 "}
    assert extract_certificate_number(row) == "С-ВЯ/15-01-2025/402123271"


def test_extract_certificate_number_missing():
    assert extract_certificate_number({}) == ""


def test_extract_certificates_from_excel_active_sheet():
    wb = Workbook()
    ws = wb.active
    ws.append([CERTIFICATE_HEADER_KEYS[1], "another column"])
    ws.append(["С-ВЯ/15-01-2025/402123271", "foo"])
    ws.append(["", "bar"])
    ws.append(["С-ВЯ/16-01-2025/402123272", "baz"])
    buf = io.BytesIO()
    wb.save(buf)

    certs = extract_certificates_from_excel(buf.getvalue())
    assert certs == [
        "С-ВЯ/15-01-2025/402123271",
        "С-ВЯ/16-01-2025/402123272",
    ]


def test_extract_certificates_from_excel_missing_header():
    wb = Workbook()
    ws = wb.active
    ws.append(["header1", "header2"])
    ws.append(["value1", "value2"])
    buf = io.BytesIO()
    wb.save(buf)

    assert extract_certificates_from_excel(buf.getvalue()) == []
